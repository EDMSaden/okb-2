import cv2 as cv
import numpy as np
import time
import win32api, win32con
import pyperclip
import openpyxl
import datetime
import os
from PIL import ImageGrab
from threading import Thread
import requests

TOKEN = "6641272022:AAFlvbJJEdex2kAzzyYGN7AxkDc2F2_tzis"
chat_id = "1036356107"
URL = 'https://api.telegram.org/bot'
#F10 - выключить все
#F2 - получить x,y
programs = set()

def get_updates(offset=0):
    result = requests.get(f'{URL}{TOKEN}/getUpdates?offset={offset}').json()
    return result['result']

def send_message(text, chat_id=1036356107):
    requests.get(f'{URL}{TOKEN}/sendMessage?chat_id={chat_id}&text={text}')

def check_message(chat_id, message):
    for mes in message.lower().replace(',', '').split():
        print(mes)
        if mes in ['/stop']:
            send_message('Выключаю бота)')
            os.system("TASKKILL /F /IM PYTHON.EXE")

def run_tg():
    update_id = get_updates()[-1]['update_id'] # Присваиваем ID последнего отправленного сообщения боту
    while True:
        time.sleep(2)
        messages = get_updates(update_id) # Получаем обновления
        for message in messages:
            # Если в обновлении есть ID больше чем ID последнего сообщения, значит пришло новое сообщение
            if update_id < message['update_id']:
                update_id = message['update_id'] # Присваиваем ID последнего отправленного сообщения боту
                # Отвечаем тому кто прислал сообщение боту
                check_message(message['message']['chat']['id'], message['message']['text'])

def stop_programs():
    while True:
        if win32api.GetKeyState(0x79) < 0:
            os.system("TASKKILL /F /IM PYTHON.EXE")

def start_program(program):
    '''Запускает программу в бесконечном цикле, принимает 'функцию def' как аргумент'''
    programs.add(program.__name__)
    exec(f'{program.__name__} = Thread(target=program, args=())')
    exec(f'{program.__name__}.start()')
    print(f'start {program.__name__}')
    send_message(f'start {program.__name__}')


def img_on_the_screen(temp):
    """Ищет изображение на экране, в качестве аргумента принемает 'название файла' в формате .bmp"""
    temp = f'{temp}.bmp'
    img_rgb = ImageGrab.grab()
    img_rgb = np.array(img_rgb)
    img_rgb = cv.cvtColor(img_rgb, cv.COLOR_BGR2RGB)
    img_gray = cv.cvtColor(img_rgb, cv.COLOR_BGR2GRAY)

    template = cv.imread(temp, cv.IMREAD_GRAYSCALE)
    if cv.imread(temp, cv.IMREAD_GRAYSCALE) is None:
        print(f"{temp} не найден, проверти наличие и формат, должен быть .bmp")
        send_message(f"{temp} не найден, проверти наличие и формат, должен быть .bmp")
        time.sleep(3)
        return False
    else:
        w,h = template.shape[::-1]
        
        res = cv.matchTemplate(img_gray, template, cv.TM_CCOEFF_NORMED)
        threshold = 0.8

        loc = np.where(res >= threshold)
        for pt in zip(*loc[::-1]):          
            if __name__ == '__main__':
                cv.rectangle(img_rgb,pt,(pt[0] + w, pt[1] + h), (0,0,255), 2)
                print(pt[0] + int(w/2), pt[1] + int(h/2))
                break
            else:
                return pt[0] + int(w/2), pt[1] + int(h/2)
            
        if __name__ == '__main__':
            img_rgb = cv.resize(img_rgb, (1024, 768))
            cv.imshow('result', img_rgb)
            cv.waitKey(0)

def wait_bars(temp):
    temp = f'{temp}.bmp'
    while True:
        time.sleep(1)
        img_rgb = ImageGrab.grab()
        img_rgb = np.array(img_rgb)
        img_gray = cv.cvtColor(img_rgb, cv.COLOR_BGR2GRAY)

        template = cv.imread(temp, cv.IMREAD_GRAYSCALE)
        assert img_rgb is not None, f"{template} не найден, проверьте наличие и формат (.bmp)"
        w,h = template.shape[::-1]
        res = cv.matchTemplate(img_gray, template, cv.TM_CCOEFF_NORMED)
        threshold = 0.8

        loc = np.where(res >= threshold)
        for pt in zip(*loc[::-1]):
            time.sleep(1)
            break

        else: break

def leftClickOn(*cord):
    """Принимает два обязательных аргумента (x,y) или один ('название изображения') в формате .bmp"""
    if type(cord[0]) is str:
        name = cord
        cord = img_on_the_screen(cord[0])
        if cord == None or cord == False:
            send_message(f'{name[0]} не найден')
            return print(f'{name[0]} не найден')
    elif len(cord) == 1:
        cord = cord[0]

    x,y = cord 
    win32api.SetCursorPos((x,y))
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,x,y,0,0)
    time.sleep(0.003)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP,x,y,0,0)
    wait_bars('wait_1')

def rightClickOn(*cord):
    """Принимает два обязательных аргумента (x,y) или один ('название изображения') в формате .bmp"""
    if type(cord[0]) is str:
        name = cord
        cord = img_on_the_screen(cord[0])
        if cord == None:
            send_message(f'{name[0]} не найден')
            return print(f'{name[0]} не найден')
    elif len(cord) == 1:
        cord = cord[0]

    x,y = cord 
    win32api.SetCursorPos((x,y))
    win32api.mouse_event(win32con.MOUSEEVENTF_RIGHTUP,x,y,0,0)
    time.sleep(0.003)
    win32api.mouse_event(win32con.MOUSEEVENTF_RIGHTDOWN,x,y,0,0)
    wait_bars('wait_1')


def doubleClickOn(*cord):
    """Принимает два обязательных аргумента (x,y) или один ('название изображения') в формате .bmp"""
    if type(cord[0]) is str:
        name = cord
        cord = img_on_the_screen(cord[0])
        if cord == None:
            send_message(f'{name[0]} не найден')
            return print(f'{name[0]} не найден')
    elif len(cord) == 1:
        cord = cord[0]

    x,y = cord
    for i in range(2):
        win32api.SetCursorPos((x,y))
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,x,y,0,0)
        time.sleep(0.003)
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP,x,y,0,0)
    wait_bars('wait_1')


keys = {'CTRL':0x11, 'ENTER':0x0D, 'PAGEUP':0x21,
        'PAGEDOWN':0x22, 'ESC':0x1B , 'BACKSPACE':0x08, 'SPACE':0x20, 'SHIFT':0x10, 'DELETE':0x2E}

def find_key(key : str):
    key = key.upper()
    if key in keys:
        key = keys.get(key)
    else: key = ord(key)
    return key

def press_key(key : str):
    """В качестве аргумента принимает любую кнопку на англ.языке ('q') или зарезервированное слово ('delete').
    Список зарезервированных слов : ctrl, enter, pageup,pagedown, esc, backspace, space, shift, delete"""
    key = find_key(key)
    win32api.keybd_event(key, 0,0,0)
    time.sleep(.05)   
    win32api.keybd_event(key,0 ,win32con.KEYEVENTF_KEYUP ,0)
    wait_bars('wait_1')

def hot_key(key_1 : str, key_2 : str):
    """Принимает два обязательных аргумента ('ctrl','a'), каждый из которых  может быть англ.буквой или зарезервированным словом
    Список зарезервированных слов : ctrl, enter, pageup,pagedown, esc, backspace, space, shift, delete.
    ⭐️Если комбинация клавиш ('ctrl' , 'c'), то возвращает значение"""
    key_1 = find_key(key_1)
    key_2 = find_key(key_2)
    win32api.keybd_event(key_1, 0,0,0)
    time.sleep(.05)
    win32api.keybd_event(key_2, 0,0,0)
    time.sleep(.05)

    win32api.keybd_event(key_1,0 ,win32con.KEYEVENTF_KEYUP ,0)
    win32api.keybd_event(key_2,0 ,win32con.KEYEVENTF_KEYUP ,0)
    wait_bars('wait_1')
    if key_1 == 0x11 and key_2 == ord('C'):
        return pyperclip.paste()


def write_text(text : str):
    """В качестве аргумента принимает ('любой текст')"""
    pyperclip.copy(text)
    hot_key('ctrl', 'v')

def change_row(x,y,text):
    """Принимает три обязательных аргумента (x, y,'любой текст')"""
    leftClickOn(x,y)
    hot_key('ctrl','a')
    write_text(f'{text}')
    press_key('enter')

class Exel:
    def __init__(self, path):
        """В качестве аргумента принимает ('название файла') в формате .xlsx"""
        self.path = path
        self.workbook = openpyxl.load_workbook(f'{path}.xlsx')
        self.work_list = self.workbook.active 
    def cell_value(self, row, colum):
        """Принимает два обязательных аргумента (поле, строка) и возвращает значение"""
        cell_value = self.work_list.cell(row = row, column = colum).value
        if type(cell_value) is datetime.datetime:
            data = self.work_list.cell(row = row, column = colum).value
            day = f'0{data.day}' if data.day < 10 else data.day
            month = f'0{data.month}' if data.month < 10 else data.month
            year = data.year
            return f'{day}.{month}.{year}'
        return cell_value
    
if __name__ != '__main__':
    th_stop_fc = Thread(target=stop_programs, args=())
    th_stop_fc.start()

    tg = Thread(target=run_tg, args=())
    tg.start()

if __name__ == '__main__':
    while True:
        if win32api.GetKeyState(0x71) < 0:
            print(f'{win32api.GetCursorPos()[0]}, {win32api.GetCursorPos()[1]}')
            pyperclip.copy(f'{win32api.GetCursorPos()[0]}, {win32api.GetCursorPos()[1]}')
            time.sleep(.5)